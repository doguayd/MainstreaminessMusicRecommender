"""
Last.fm 360K  →  dataset_lfm.xlsx
==================================
Reads the real Last.fm 360K dataset and produces an Excel file whose
sheet structure is identical to dataset_large.xlsx so that
Mainstream_final.py / mainstream_big.py can load it without any changes.

Output sheets
-------------
  Playcounts  – user × artist playcount matrix  (header on row 3, i.e. header=2)
  Users       – user_id, country_code, country_name, gender, age
  Artists     – artist name + metadata
  Countries   – country summary statistics

Target: 500 users × ~81 artists × 12 countries  (same scale as synthetic dataset)
"""

import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE      = os.path.dirname(os.path.abspath(__file__))
LFM_DIR   = os.path.join(BASE, "lastfm-dataset-360K", "lastfm-dataset-360K")
PLAYS_TSV = os.path.join(LFM_DIR, "usersha1-artmbid-artname-plays.tsv")
PROF_TSV  = os.path.join(LFM_DIR, "usersha1-profile.tsv")
OUT_PATH  = os.path.join(BASE, "fourth progress", "dataset_lfm.xlsx")

# ── Country mapping (Last.fm name → ISO code) ────────────────────────────────
COUNTRY_MAP = {
    "United States":        ("US", "United States"),
    "Germany":              ("DE", "Germany"),
    "United Kingdom":       ("GB", "United Kingdom"),
    "Japan":                ("JP", "Japan"),
    "Turkey":               ("TR", "Turkey"),
    "Brazil":               ("BR", "Brazil"),
    "France":               ("FR", "France"),
    "Korea, Republic of":   ("KR", "South Korea"),
    "Finland":              ("FI", "Finland"),
    "Sweden":               ("SE", "Sweden"),
    "Italy":                ("IT", "Italy"),
    "Mexico":               ("MX", "Mexico"),
}

# ── Per-country user quotas (matching synthetic dataset proportions) ──────────
QUOTA = {
    "US": 79, "DE": 50, "GB": 50, "JP": 45,
    "TR": 40, "BR": 40, "FR": 36, "KR": 25,
    "FI": 31, "SE": 31, "IT": 31, "MX": 31,
}
TARGET_ARTISTS = 81   # how many artists to keep in the matrix

# ─────────────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 1 — Loading user profiles …")
print("=" * 70)

profile = pd.read_csv(
    PROF_TSV, sep="\t", header=None,
    names=["user", "gender", "age", "country", "signup"],
    dtype={"user": str, "gender": str, "age": str, "country": str, "signup": str},
)
profile["age"] = pd.to_numeric(profile["age"], errors="coerce")

# ── Sample users per country ──────────────────────────────────────────────────
selected_users = []
selected_meta  = []

np.random.seed(42)
for country_full, (code, name) in COUNTRY_MAP.items():
    pool = profile[profile["country"] == country_full].copy()
    # Prefer users with valid gender/age entries for richer metadata
    pool_complete = pool[pool["gender"].isin(["m", "f"]) & pool["age"].notna()]
    quota = QUOTA[code]

    if len(pool_complete) >= quota:
        chosen = pool_complete.sample(quota, random_state=42)
    elif len(pool) >= quota:
        chosen = pool.sample(quota, random_state=42)
    else:
        chosen = pool

    for _, row in chosen.iterrows():
        selected_users.append(row["user"])
        selected_meta.append({
            "user_id":       row["user"],
            "country_code":  code,
            "country_name":  name,
            "gender":        row["gender"] if pd.notna(row["gender"]) else "?",
            "age":           int(row["age"]) if pd.notna(row["age"]) else None,
        })

selected_set = set(selected_users)
print(f"  Selected {len(selected_users)} users across {len(QUOTA)} countries.")

# ─────────────────────────────────────────────────────────────────────────────
print("\nSTEP 2 — Scanning plays file for selected users …")
print("  (17.5 M rows — this takes ~30 s)")
print("=" * 70)

CHUNK = 500_000
play_rows = []

for i, chunk in enumerate(
    pd.read_csv(
        PLAYS_TSV, sep="\t", header=None,
        names=["user", "artist_id", "artist_name", "plays"],
        dtype={"user": str, "artist_id": str, "artist_name": str, "plays": str},
        chunksize=CHUNK,
    )
):
    chunk["plays"] = pd.to_numeric(chunk["plays"], errors="coerce").fillna(0).astype(int)
    sub = chunk[chunk["user"].isin(selected_set)]
    play_rows.append(sub)
    if (i + 1) % 5 == 0:
        print(f"  … processed {(i+1)*CHUNK:,} rows")

plays = pd.concat(play_rows, ignore_index=True)
print(f"  Filtered down to {len(plays):,} play records.")

# ─────────────────────────────────────────────────────────────────────────────
print("\nSTEP 3 — Selecting top artists …")
print("=" * 70)

# Normalise artist names (lowercase, strip whitespace)
plays["artist_name"] = plays["artist_name"].str.strip().str.lower()

# Pick top artists by total play count across all selected users
artist_totals = plays.groupby("artist_name")["plays"].sum().sort_values(ascending=False)
# Require at least 5 users to have listened (avoid very niche artists)
artist_listeners = plays.groupby("artist_name")["user"].nunique()
eligible = artist_totals[artist_listeners >= 5]
top_artists = eligible.head(TARGET_ARTISTS).index.tolist()

print(f"  Top {len(top_artists)} artists selected.")
print(f"  #1: {top_artists[0]}   #5: {top_artists[4]}   #{TARGET_ARTISTS}: {top_artists[-1]}")

# ─────────────────────────────────────────────────────────────────────────────
print("\nSTEP 4 — Building user × artist matrix …")
print("=" * 70)

plays_top = plays[plays["artist_name"].isin(top_artists)]
matrix    = plays_top.pivot_table(
    index="user", columns="artist_name", values="plays", aggfunc="sum", fill_value=0
)
# Reindex to include all selected users and all top artists (fill 0 where missing)
matrix = matrix.reindex(index=selected_users, columns=top_artists, fill_value=0)
print(f"  Matrix shape: {matrix.shape[0]} users × {matrix.shape[1]} artists")

# ─────────────────────────────────────────────────────────────────────────────
print("\nSTEP 5 — Assembling DataFrames …")
print("=" * 70)

meta_df = pd.DataFrame(selected_meta).set_index("user_id")

# Playcounts sheet DataFrame (matches existing format)
play_df = matrix.copy()
play_df.insert(0, "Country Name", [meta_df.loc[u, "country_name"] for u in play_df.index])
play_df.insert(0, "Country\nCode", [meta_df.loc[u, "country_code"] for u in play_df.index])
play_df.index.name = "User ID"

# Artists metadata
artist_col_display = [a.title() for a in top_artists]   # Title-case for display
artist_totals_top  = plays_top.groupby("artist_name")["plays"].sum().reindex(top_artists)
artist_listeners_top = plays_top.groupby("artist_name")["user"].nunique().reindex(top_artists)

# Country-specific classification (simple heuristic based on listening concentration)
country_concentration = {}
for art in top_artists:
    art_plays = plays_top[plays_top["artist_name"] == art].copy()
    art_plays["country"] = art_plays["user"].map(meta_df["country_code"])
    total = art_plays["plays"].sum()
    if total == 0:
        country_concentration[art] = "Global"
        continue
    top_c = art_plays.groupby("country")["plays"].sum().idxmax()
    top_frac = art_plays.groupby("country")["plays"].sum().max() / total
    country_concentration[art] = top_c if top_frac > 0.5 else "Global"

artists_df = pd.DataFrame({
    "Artist":             [a.title() for a in top_artists],
    "Type":               [("Country-Specific" if country_concentration[a] != "Global" else "Global") for a in top_artists],
    "Dominant Country":   [country_concentration[a] for a in top_artists],
    "Total Plays":        [int(artist_totals_top[a]) for a in top_artists],
    "Listener Count":     [int(artist_listeners_top[a]) for a in top_artists],
    "Source":             "Last.fm 360K (real data)",
})

# Users sheet
users_df = meta_df.reset_index()
users_df.columns = ["User ID", "Country Code", "Country Name", "Gender", "Age"]
users_df["Data Source"] = "Last.fm 360K"

# Countries sheet
country_stats = []
for lfm_name, (code, name) in COUNTRY_MAP.items():
    n_users = sum(1 for u in selected_users if meta_df.loc[u, "country_code"] == code)
    country_stats.append({
        "Code": code, "Country": name,
        "Users Selected": n_users,
        "LFM 360K Total": len(profile[profile["country"] == lfm_name]),
        "Notes": f"Real Last.fm listeners. Quota: {QUOTA[code]}",
    })
countries_df = pd.DataFrame(country_stats)

# ─────────────────────────────────────────────────────────────────────────────
print("\nSTEP 6 — Writing Excel workbook …")
print("=" * 70)

wb = openpyxl.Workbook()

# ── Helper styles ─────────────────────────────────────────────────────────────
def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT   = Font(bold=True, color="FFFFFF", size=12)
MONO_FONT    = Font(name="Courier New", size=9)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left", vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Country color palette (same logic as synthetic dataset)
COUNTRY_COLORS = {
    "US": "DCE6F1", "DE": "E2EFDA", "GB": "FFF2CC", "JP": "FCE4D6",
    "TR": "F4CCCC", "BR": "D9EAD3", "FR": "CFE2F3", "KR": "EAD1DC",
    "FI": "D9D2E9", "SE": "FFF9C4", "IT": "F4CCCC", "MX": "FCE5CD",
}
HEADER_COLOR = "2F5496"
ZERO_COLOR   = "F2F2F2"


# ─── Sheet 1: Playcounts ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Playcounts"

# Row 1: title banner
ws1.merge_cells(f"A1:{get_column_letter(3 + len(top_artists))}1")
title_cell = ws1["A1"]
title_cell.value = (
    f"Last.fm 360K — Real User Playcount Data  |  "
    f"{len(selected_users)} users × {len(top_artists)} artists × {len(QUOTA)} countries"
)
title_cell.font      = TITLE_FONT
title_cell.fill      = header_fill(HEADER_COLOR)
title_cell.alignment = CENTER
ws1.row_dimensions[1].height = 22

# Row 2: dataset info
ws1.merge_cells(f"A2:{get_column_letter(3 + len(top_artists))}2")
info_cell = ws1["A2"]
info_cell.value = (
    "Source: Last.fm Dataset 360K (Celma, O. 2010).  "
    "Subset selected proportionally from 12 countries.  "
    "Plays: raw listen counts as recorded on Last.fm."
)
info_cell.font      = Font(italic=True, size=9, color="595959")
info_cell.alignment = CENTER
ws1.row_dimensions[2].height = 16

# Row 3: column headers
headers = ["User ID", "Country\nCode", "Country Name"] + [a.title() for a in top_artists]
for col_idx, h in enumerate(headers, 1):
    c = ws1.cell(row=3, column=col_idx, value=h)
    c.font      = HEADER_FONT
    c.fill      = header_fill(HEADER_COLOR)
    c.alignment = CENTER
    c.border    = THIN_BORDER
ws1.row_dimensions[3].height = 30

# Freeze panes
ws1.freeze_panes = "D4"

# Data rows
for row_idx, uid in enumerate(selected_users, 4):
    code = meta_df.loc[uid, "country_code"]
    name = meta_df.loc[uid, "country_name"]
    row_fill = PatternFill("solid", fgColor=COUNTRY_COLORS.get(code, "FFFFFF"))

    ws1.cell(row=row_idx, column=1, value=uid[:20]).font = Font(name="Courier New", size=8)
    ws1.cell(row=row_idx, column=1).fill = row_fill
    ws1.cell(row=row_idx, column=1).alignment = LEFT

    ws1.cell(row=row_idx, column=2, value=code).fill = row_fill
    ws1.cell(row=row_idx, column=2).alignment = CENTER
    ws1.cell(row=row_idx, column=2).font = Font(bold=True, size=9)

    ws1.cell(row=row_idx, column=3, value=name).fill = row_fill
    ws1.cell(row=row_idx, column=3).alignment = LEFT

    for col_idx, artist in enumerate(top_artists, 4):
        val = int(matrix.loc[uid, artist])
        c   = ws1.cell(row=row_idx, column=col_idx, value=val if val > 0 else 0)
        c.fill      = PatternFill("solid", fgColor=ZERO_COLOR) if val == 0 else row_fill
        c.alignment = CENTER
        c.font      = Font(size=8, color="999999" if val == 0 else "000000")

# Column widths
ws1.column_dimensions["A"].width = 24
ws1.column_dimensions["B"].width = 8
ws1.column_dimensions["C"].width = 16
for i in range(4, 4 + len(top_artists)):
    ws1.column_dimensions[get_column_letter(i)].width = 12


# ─── Sheet 2: Users ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Users")
ws2.append(["User ID", "Country Code", "Country Name", "Gender", "Age", "Data Source"])
for cell in ws2[1]:
    cell.font = HEADER_FONT
    cell.fill = header_fill(HEADER_COLOR)
    cell.alignment = CENTER

for _, row in users_df.iterrows():
    ws2.append(list(row))

ws2.column_dimensions["A"].width = 44
for col in ["B","C","D","E","F"]:
    ws2.column_dimensions[col].width = 16
ws2.freeze_panes = "A2"


# ─── Sheet 3: Artists ────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Artists")
ws3.append([""] * len(artists_df.columns))   # blank row (matches header=1 in pd.read_excel)
ws3.append(list(artists_df.columns))
for cell in ws3[2]:
    cell.font = HEADER_FONT
    cell.fill = header_fill(HEADER_COLOR)
    cell.alignment = CENTER

for _, row in artists_df.iterrows():
    ws3.append(list(row))

for col in ["A","B","C","D","E","F"]:
    ws3.column_dimensions[col].width = 22
ws3.freeze_panes = "A3"


# ─── Sheet 4: Countries ──────────────────────────────────────────────────────
ws4 = wb.create_sheet("Countries")
ws4.append(list(countries_df.columns))
for cell in ws4[1]:
    cell.font = HEADER_FONT
    cell.fill = header_fill(HEADER_COLOR)
    cell.alignment = CENTER

for _, row in countries_df.iterrows():
    ws4.append(list(row))

for col in ["A","B","C","D","E"]:
    ws4.column_dimensions[col].width = 20


# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"\n✓ Saved: {OUT_PATH}")
print(f"  Playcounts sheet : {len(selected_users)} rows × {len(top_artists)+3} cols")
print(f"  Users sheet      : {len(users_df)} rows")
print(f"  Artists sheet    : {len(artists_df)} rows")
print(f"  Countries sheet  : {len(countries_df)} rows")
print("\nDone!")
